from random import randint
from openpyxl import Workbook, load_workbook

# membuka file excel
wb = Workbook()

# membaca worksheet
ws = wb.active

# membaca cell
# perhitungan row dan column dimulai dari 1 (bukan 0)
ws.cell(row=1, column=1).value = randint(10,100)


n = 1
while n<=5:
    wb.save("file"+str(n)+".xlsx")
    n += 1