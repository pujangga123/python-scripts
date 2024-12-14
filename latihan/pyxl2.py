from random import randint
from openpyxl import Workbook, load_workbook

# membuat objek workbook
wb = Workbook()

# set sheet
ws = wb.active

# isi data ke sheet
ws.cell(row=1, column=1).value = "100 data acak"
# input 100 bilangan acak antara 1-100
n = 1
while n<=100:
    ws.cell(row=2+n, column=1).value = randint(1,100)
    n += 1

# simpan file
wb.save("contoh1.xlsx")